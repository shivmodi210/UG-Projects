from pulp import *
import openpyxl as xl
import sys
import solver
food_items={}
wb=xl.load_workbook('Mess_menu.xlsx')
sheet_nutri=wb['Macronutrients']

data_row=0
while(True):
    data_row+=1
    if type(sheet_nutri.cell(data_row,2).value)==float:
        break
for i in range(data_row,sheet_nutri.max_row+1):
    
    if type(sheet_nutri.cell(i,2).value)!=float and type(sheet_nutri.cell(i,2).value)!=int:
        break
    
    food_name=sheet_nutri.cell(i,1).value
    food_items[food_name]=[]
    for j in range(2,7):
        cell=sheet_nutri.cell(i,j)    
        food_items[food_name].append(cell.value)   

sheet_menu=wb['Menu']
sheet_pref=wb["User_Input"]
# user data

weight=float(input("Enter your weight (kg):"))
height=int(input("Enter your height (cm) :"))
age=int(input("Enter your age :"))
gender=input("Enter your gender (M/F):")

if gender=='M':
    bmr=10*weight+6.25*height-5*age+5
elif gender=="F":
    bmr=10*weight+6.25*height-5*age-161
else:
    print("Please select gender from the given options.")
    sys.exit()
print("S: Sitting all day with no structured excercise.")
print("M: Moderate movement or 1hr exercise")
print("H: 2hrs exercise or moderate movement with 1hr exercise")
print("E: High training or high movement with 1 +hrs of exercise.")
pa=input("Input physical activity level (S/M/H/E):")                #physical activity
exer_cal=int(input("Enter Exercise Calories:"))
phy_act_lev={
    "S":1.55,"M":1.85,"H":2.2,"E":2.4}
par=phy_act_lev.get(pa)
if par!=None:
    maint_cal=bmr*par
else:
    print("Please select physical Activity from the given options.")
    sys.exit()
    
goal=input("Target (G/M/L):")
gain=int(input("Input targetted calorie deficiet or surplus:"))
CalPD=maint_cal+gain

if goal=="G":
    food_items["None"][-1]=20
else:
    food_items["None"][-1]=-20
# weights solver
pref=[]
for i in range(3,sheet_pref.max_row+1):
    if type(sheet_pref.cell(i,2).value)!=int:
            break
    pref.append([])
    for j in range(2,sheet_pref.max_column+1):
        if type(sheet_pref.cell(i,j).value)!=int:
            break
        pref[i-3].append(sheet_pref.cell(i,j).value)
        
food_item_weights={}
food_item_weights2={}
food_menu=[]
for i in range(1,sheet_menu.max_row+1):
    if type(sheet_menu.cell(i,2).value)!=str:
        break
    food_menu.append([])
    for j in range(2,sheet_menu.max_column+1):
        if type(sheet_menu.cell(i,j).value)!=str:
            break
        food_menu[i-1].append(sheet_menu.cell(i,j).value)
        if i>=3:
            food_name=sheet_menu.cell(i,j).value+","+sheet_menu.cell(2,j).value+","+sheet_menu.cell(1,j).value
            food_name2=sheet_menu.cell(i,j).value+","+sheet_menu.cell(2,j).value
            food_item_weights[food_name]=1
            food_item_weights2[food_name2]=1
   
food_item_weights_user=solver.preferance(pref,food_item_weights,food_menu)
#lp solver
data_row=0
while(True):
    data_row+=1
    if sheet_menu.cell(data_row,1).value=="Monday":
        break

for day in range(7):   
    day_menu=[]
    for i in range(2,sheet_menu.max_column+1):
        if type(sheet_menu.cell(data_row+day,i).value)!=str:
            break
        day_menu.append(sheet_menu.cell(data_row+day,i).value+","+sheet_menu.cell(2,i).value)
    day_meal_weights=[]
    for i in range(len(day_menu)):
        day_meal_weights.append(food_item_weights2[day_menu[i]])
    meal_items_solver=[]
    for i in range(len(day_menu)):
        meal_items_solver.append([])
        meal_items_solver[i].append(day_menu[i])
        meal_items_solver[i].append(food_items.get((day_menu[i].split(","))[0]))
    quantity=solver.mess_plan(meal_items_solver,day_meal_weights,weight,goal,CalPD)
    input_cell=sheet_menu.cell(data_row+7+day,1)
    input_cell.value=sheet_menu.cell(data_row+day,1).value
    for i in range(2,sheet_menu.max_column+1):
        if type(sheet_menu.cell(data_row+day,i).value)!=str:
            break
        input_cell=sheet_menu.cell(data_row+7+day,i)
        input_cell.value=quantity[i-2][-1]
wb.save("Output.xlsx")


# output after analyzing preferance
data_row=0
while(True):
    data_row+=1
    if sheet_menu.cell(data_row,1).value=="Monday":
        break
    
input_cell=sheet_menu.cell(data_row+14,1,1)
input_cell.value="Mess Plan after analysing user data"

for day in range(7):   
    day_menu=[]
    for i in range(2,sheet_menu.max_column+1):
        if type(sheet_menu.cell(data_row+day,i).value)!=str:
            break
        day_menu.append(sheet_menu.cell(data_row+day,i).value+","+sheet_menu.cell(2,i).value)
    day_meal_weights=[]
    for i in range(len(day_menu)):
        day_meal_weights.append(food_item_weights_user[day_menu[i]])
#        day_meal_weights.append(food_item_weights_user[day_menu[i]]*food_items[day_menu[i].split(",")[0]][-1]) for 3rd solution
    meal_items_solver=[]
    for i in range(len(day_menu)):
        meal_items_solver.append([])
        meal_items_solver[i].append(day_menu[i])
        meal_items_solver[i].append(food_items.get((day_menu[i].split(","))[0]))
    quantity=solver.mess_plan(meal_items_solver,day_meal_weights,weight,goal,CalPD)
    input_cell=sheet_menu.cell(data_row+15+day,1)
    input_cell.value=sheet_menu.cell(data_row+day,1).value
    for i in range(2,sheet_menu.max_column+1):
        if type(sheet_menu.cell(data_row+day,i).value)!=str:
            break
        input_cell=sheet_menu.cell(data_row+15+day,i)
        input_cell.value=quantity[i-2][-1]
wb.save("Output_1st_method.xlsx")

